import sys
import tkinter
from tkinter import*
from tkinter import messagebox
import datetime
import json
from tkinter.filedialog import asksaveasfile
import pprint
from openpyxl import Workbook,load_workbook
import os


try:
    import Tkinter as tk
except ImportError:
    import tkinter as tk

try:
    import ttk
    py3 = False
except ImportError:
    import tkinter.ttk as ttk

    py3 = True

import Patient_Info_support

def vp_start_gui():
    '''Starting point when module is the main routine.'''
    global val, w, root
    root = tk.Tk()
    Patient_Info_support.set_Tk_var()
    top = Toplevel1 (root)
    Patient_Info_support.init(root, top)
    #root.overrideredirect(2)
    root.mainloop()

w = None
def create_Toplevel1(rt, *args, **kwargs):
    '''Starting point when module is imported by another module.
       Correct form of call: 'create_Toplevel1(root, *args, **kwargs)' .'''
    global w, w_win, root
    #rt = root
    root = rt
    w = tk.Toplevel (root)
    Patient_Info_support.set_Tk_var()
    top = Toplevel1 (w)
    Patient_Info_support.init(w, top, *args, **kwargs)
    return (w, top)

def destroy_Toplevel1():
    global w
    w.destroy()
    w = None

def character_limit(entry_text):
    if len(entry_text.get()) > 0:
        entry_text.set(entry_text.get()[:2])
        

def character_limit5(entry_text5):
    if len(entry_text5.get()) > 0:
        entry_text5.set(entry_text5.get()[:2])
        
            

def character_limit1(entry_text1):
    if len(entry_text1.get()) > 0:
        entry_text1.set(entry_text1.get()[:10])
        
def character_limit2(entry_text2):
    if len(entry_text2.get()) > 0:
        entry_text2.set(entry_text2.get()[:12])

def character_limit3(entry_text3):
    if len(entry_text3.get()) > 0:
        entry_text3.set(entry_text3.get()[:12])
    




def testVal(inStr,acttyp):
    if acttyp == '1': 
        if not inStr.isdigit():   
            return False
    return True
    
def testVal1(inStr,acttyp):
    if acttyp == '1': 
        if not inStr.isalpha():   
            return False
    return True





     

def cancel():
        if messagebox.askokcancel("Exit", "Are you sure?"):
                root.destroy()


def writeToJSONFile(path, fileName, data):
        json.dump(data, path)
 
path = "./home/pi/Maternal_care_tk"

#def writeToJSONFile(path, fileName, data):
    #filePathNameWExt = './path/Data' + fileName + '.json'

class Toplevel1:
    #def male(self):
        #e=str(self.var.get())
        #if '1' in e:
            #e = "Male"
            #print(e)
       
    #def female(self):
        #e=str(self.var.get())
        #if '2' in e:
            #e = "Female"
            #print(e)
    #def other(self):
         #e=str(self.var.get())
         #if '3' in e:
            #e = "Other"
            #print(e)

    def next_step(self):
        if self.Entry1_2.get():
            #if self.Entry1_3.get():
        # the user entered data in the mandatory entry: proceed to next step
            print("next step")
                #root.destroy()
        else:
        # the mandatory field is empty
            print("mandatory data missing")
            self.Entry1_2.focus_set()
                #self.Entry1_3.focus_set()


    #def writeToJSONFile(path, fileName, data):
        #json.dump(data, path)
        
 
    #path = './home/pi/Maternal_care_tk/Data'
    #test = 1
    
    
    
 
 
    def submit(self):  
        a = self.Entry1.get()
        a = a.zfill(12)
        b = self.Entry1_1.get()
        c = self.Entry1_2.get()
        d = self.Entry1_3.get()
        print(a)
        print(b)
        print(c)
        print(d)
        e=str(self.var.get())
        if '1' in e:
            e = "Male"
            print(e)
            
        elif '2' in e:
            e = "Female"
            print(e)
        
        elif '3' in e:
            e = "Other"
            print(e)
            
        f = self.Entry1_4.get()
        g = self.Entry1_5.get()
        h = self.Entry1_6.get()
        
        
        i = self.Entry1_7.get()
        j = self.Entry1_8.get()
        
        print(f)
        print(g)
        print(h)
        
        print(i)
        print(j)
        
        #data to save         
        patient_data=[i,a,b,c,d,e,f,g,h,j]
        
        #File path for excel
        excel_file_path = "/path/patient_data.xlsx"
        
        if os.path.exists(excel_file_path):
            
        #Load the existing workbook
            workbook =load_workbook(excel_file_path)
            sheet =workbook.active
            
        else:
            
        # created a new workbook and sheet
            workbook =Workbook()
            sheet =workbook.active
            
#Add headers if creating a new file
            headers=["Bed No","Patient ID","Date","Patient name","Age","Gender","Mobile","Aadhar","Address","Finger ID"]
            sheet.append(headers)
        
    #     Append the new Data
        sheet.append(patient_data)
        
    #     Save the workbook
        workbook.save(excel_file_path)
        
        messagebox.showinfo("Success","Patient data added to excel successfully!")
        
        
        
        

        if(self.Entry1_2.get() == ""):
            messagebox.showerror("Error", "please fill mandatory field")
        elif(self.Entry1_3.get() == ""):
            messagebox.showerror("Error", "please fill mandatory field")
        elif int(f) <= 1000000000:
            messagebox.showerror("Error", "Please insert correct Mobile Number")
        elif int(g) <= 100000000000:
            messagebox.showerror("Error", "Please insert correct Aadhar Number")
        elif self.Entry1_2.get().isdigit() and not self.Entry1_2.get().isalpha():
            messagebox.showerror('Only letters','Please enter your name in string only')
        #elif self.Entry1_2.get().isalpha():
            #messagebox.showerror('Only letters','Please enter your name in string only')
#             root.destroy()
        else:
            root.destroy()
            try:
                command = f"python /home/pi/Project/esp32_send.py {i} {a} {c} {j}"
                os.system(command)
                messagebox.showinfo("Success","Other script executed successfully!")
                root.destroy()
                
            except Exception as e:
                messagebox.showerror("Error",f"Error excecuting script: {e}!")
                

      
    def submit_old(self):  
        a = self.Entry1.get()
        a = a.zfill(12)
        b = self.Entry1_1.get()
        c = self.Entry1_2.get()
        d = self.Entry1_3.get()
        print(a)
        print(b)
        print(c)
        print(d)
        e=str(self.var.get())
        if '1' in e:
            e = "Male"
            print(e)
            
        elif '2' in e:
            e = "Female"
            print(e)
        
        elif '3' in e:
            e = "Other"
            print(e)
            
        f = self.Entry1_4.get()
        g = self.Entry1_5.get()
        h = self.Entry1_6.get()
        
        
        i = self.Entry1_7.get()
        j = self.Entry1_8.get()
        
        print(f)
        print(g)
        print(h)
        
        print(i)
        print(j)
        data = {}
        data['Patient Id'] = a
        data['Date'] = b
        data['Patient Name'] = c
        data['Age'] = d
        data['Gender'] = e
        data['Mobile'] = f
        data['Aadhar'] = g
        data['Address'] = h
        
        data['Bed No'] = i
        data['Finger ID'] = g
        file = open('/home/pi/Project/jsondata' + '.json', "w")
        files = [('JSON File', '*.json')]
        fileName='IOTEDU'
        writeToJSONFile(file, fileName, data)
#         with open("/etc/wpa_supplicant/wpa_supplicant.conf") as file:
#                 data1 = file.read()
#                 print(data1)
        #filepos = asksaveasfile(filetypes = files,defaultextension = json,initialfile= "jsondata" + '.json')
        #writeToJSONFile(filepos, fileName, data)
        #print("Patient Id: %s\nDate: %s\nPatient Name: %s\nage: %s\nGender: %s\nMobile: %s\nAadhar: %s\nAddress: %s" % (self.Entry1.get(), self.Entry1_1.get(), self.Entry1_2.get(), self.Entry1_3.get(), self.Entry1_4.get(), self.Entry1_5.get(), self.Entry1_6.get()))
        #self.Entry1.delete(0,END)
        #self.Entry1_1.delete(0,END)
        #self.Entry1_2.delete(0,END)
        #self.Entry1_3.delete(0,END)
        #self.Entry1_4.delete(0,END)
        #self.Entry1_5.delete(0,END)
        #self.Entry1_6.delete(0,END)
        if(self.Entry1_2.get() == ""):
            messagebox.showerror("Error", "please fill mandatory field")
        elif(self.Entry1_3.get() == ""):
            messagebox.showerror("Error", "please fill mandatory field")
        elif int(f) <= 1000000000:
            messagebox.showerror("Error", "Please insert correct Mobile Number")
        elif int(g) <= 100000000000:
            messagebox.showerror("Error", "Please insert correct Aadhar Number")
        elif self.Entry1_2.get().isdigit() and not self.Entry1_2.get().isalpha():
            messagebox.showerror('Only letters','Please enter your name in string only')
        #elif self.Entry1_2.get().isalpha():
            #messagebox.showerror('Only letters','Please enter your name in string only')
        else:
             root.destroy()

    

        


    def __init__(self, top=None):
        '''This class configures and populates the toplevel window.
           top is the toplevel containing window.'''
        _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
        _fgcolor = '#000000'  # X11 color: 'black'
        _compcolor = '#d9d9d9' # X11 color: 'gray85'
        _ana1color = '#d9d9d9' # X11 color: 'gray85'
        _ana2color = '#ececec' # Closest X11 color: 'gray92'
        
        entry_text = StringVar()
        entry_text.trace("w", lambda *args: character_limit(entry_text))
        entry_text1 = StringVar()
        entry_text1.trace("w", lambda *args: character_limit1(entry_text1))
        entry_text2 = StringVar()
        entry_text2.trace("w", lambda *args: character_limit2(entry_text2))
        entry_text3 = StringVar()
        entry_text3.trace("w", lambda *args: character_limit3(entry_text3))
        entry_text4 = StringVar()
        entry_text4.trace("w", lambda *args: character_limit3(entry_text3))
        
        entry_text5 = StringVar()
        entry_text5.trace("w", lambda *args: character_limit5(entry_text5))

       # sort_value = IntVar()

        
        top.geometry("400x450+550+76")
        top.minsize(120, 1)
        top.maxsize(1370, 749)
        top.resizable(0,  0)
        top.title("New Toplevel")
        top.configure(background="#d9d9d9")
        top.configure(highlightbackground="#d9d9d9")
        top.configure(highlightcolor="black")

        
        self.var=IntVar()
        self.Frame1 = tk.Frame(top)
        self.Frame1.place(relx=0.05, rely=0.044, relheight=0.911
                , relwidth=0.915)
        self.Frame1.configure(relief='groove')
        self.Frame1.configure(borderwidth="2")
        self.Frame1.configure(relief="groove")
        self.Frame1.configure(background="#d9d9d9")
        self.Frame1.configure(highlightbackground="#d9d9d9")
        self.Frame1.configure(highlightcolor="black")

        self.Label1 = tk.Label(self.Frame1)
        self.Label1.place(relx=0.071, rely=0.278, height=28, width=104)
        self.Label1.configure(activebackground="#f9f9f9")
        self.Label1.configure(activeforeground="black")
        self.Label1.configure(background="#d9d9d9")
        self.Label1.configure(disabledforeground="#a3a3a3")
        self.Label1.configure(foreground="#000000")
        self.Label1.configure(highlightbackground="#d9d9d9")
        self.Label1.configure(highlightcolor="black")
        self.Label1.configure(text='''Patient Name''')

        self.Label2 = tk.Label(self.Frame1)
        self.Label2.place(relx=0.018, rely=0.369, height=18, width=81)
        self.Label2.configure(activebackground="#f9f9f9")
        self.Label2.configure(activeforeground="black")
        self.Label2.configure(background="#d9d9d9")
        self.Label2.configure(disabledforeground="#a3a3a3")
        self.Label2.configure(foreground="#000000")
        self.Label2.configure(highlightbackground="#d9d9d9")
        self.Label2.configure(highlightcolor="black")
        self.Label2.configure(text='''Age''')

        self.Label3 = tk.Label(self.Frame1)
        self.Label3.place(relx=0.066, rely=0.436, height=28, width=70)
        self.Label3.configure(activebackground="#f9f9f9")
        self.Label3.configure(activeforeground="black")
        self.Label3.configure(background="#d9d9d9")
        self.Label3.configure(disabledforeground="#a3a3a3")
        self.Label3.configure(foreground="#000000")
        self.Label3.configure(highlightbackground="#d9d9d9")
        self.Label3.configure(highlightcolor="black")
        self.Label3.configure(text='''Gender''')

        self.Label4 = tk.Label(self.Frame1)
        self.Label4.place(relx=0.066, rely=0.506, height=36, width=70)
        self.Label4.configure(activebackground="#f9f9f9")
        self.Label4.configure(activeforeground="black")
        self.Label4.configure(background="#d9d9d9")
        self.Label4.configure(disabledforeground="#a3a3a3")
        self.Label4.configure(foreground="#000000")
        self.Label4.configure(highlightbackground="#d9d9d9")
        self.Label4.configure(highlightcolor="black")
        self.Label4.configure(text='''Mobile''')

        self.Label5 = tk.Label(self.Frame1)
        self.Label5.place(relx=0.073, rely=0.597, height=28, width=70)
        self.Label5.configure(activebackground="#f9f9f9")
        self.Label5.configure(activeforeground="black")
        self.Label5.configure(background="#d9d9d9")
        self.Label5.configure(disabledforeground="#a3a3a3")
        self.Label5.configure(foreground="#000000")
        self.Label5.configure(highlightbackground="#d9d9d9")
        self.Label5.configure(highlightcolor="black")
        self.Label5.configure(text='''Aadhar''')

        self.Label6 = tk.Label(self.Frame1)
        self.Label6.place(relx=0.082, rely=0.668, height=36, width=70)
        self.Label6.configure(activebackground="#f9f9f9")
        self.Label6.configure(activeforeground="black")
        self.Label6.configure(background="#d9d9d9")
        self.Label6.configure(disabledforeground="#a3a3a3")
        self.Label6.configure(foreground="#000000")
        self.Label6.configure(highlightbackground="#d9d9d9")
        self.Label6.configure(highlightcolor="black")
        self.Label6.configure(text='''Address''')
        
        
        
        self.Label7 = tk.Label(self.Frame1)
        self.Label7.place(relx=0.062, rely=0.100, height=17, width=70)
        self.Label7.configure(activebackground="#f9f9f9")
        self.Label7.configure(activeforeground="black")
        self.Label7.configure(background="#d9d9d9")
        self.Label7.configure(disabledforeground="#a3a3a3")
        self.Label7.configure(foreground="#000000")
        self.Label7.configure(highlightbackground="#d9d9d9")
        self.Label7.configure(highlightcolor="black")
        self.Label7.configure(text='''Bed No.''')
        
                
        self.Label8 = tk.Label(self.Frame1)
        self.Label8.place(relx=0.062, rely=0.760, height=17, width=70)
        self.Label8.configure(activebackground="#f9f9f9")
        self.Label8.configure(activeforeground="black")
        self.Label8.configure(background="#d9d9d9")
        self.Label8.configure(disabledforeground="#a3a3a3")
        self.Label8.configure(foreground="#000000")
        self.Label8.configure(highlightbackground="#d9d9d9")
        self.Label8.configure(highlightcolor="black")
        self.Label8.configure(text='''FG No.''')
        

        self.Label13 = tk.Label(self.Frame1)
        self.Label13.place(relx=0.059, rely=0.156, height=17, width=82)
        self.Label13.configure(activebackground="#f9f9f9")
        self.Label13.configure(activeforeground="black")
        self.Label13.configure(background="#d9d9d9")
        self.Label13.configure(disabledforeground="#a3a3a3")
        self.Label13.configure(foreground="#000000")
        self.Label13.configure(highlightbackground="#d9d9d9")
        self.Label13.configure(highlightcolor="black")
        self.Label13.configure(text='''Patient ID''')

        self.Label14 = tk.Label(self.Frame1)
        self.Label14.place(relx=0.051, rely=0.208, height=27, width=59)
        self.Label14.configure(activebackground="#f9f9f9")
        self.Label14.configure(activeforeground="black")
        self.Label14.configure(background="#d9d9d9")
        self.Label14.configure(disabledforeground="#a3a3a3")
        self.Label14.configure(foreground="#000000")
        self.Label14.configure(highlightbackground="#d9d9d9")
        self.Label14.configure(highlightcolor="black")
        self.Label14.configure(text='''Date''')

        self.Label15 = tk.Label(self.Frame1)
        self.Label15.place(relx=0.003, rely=-0.003, height=28, width=331)
        self.Label15.configure(activebackground="#f9f9f9")
        self.Label15.configure(activeforeground="black")
        self.Label15.configure(background="#80ffff")
        self.Label15.configure(disabledforeground="#a3a3a3")
        self.Label15.configure(font="-family {Segoe UI} -size 14 -weight bold")
        self.Label15.configure(foreground="#000000")
        self.Label15.configure(highlightbackground="#d9d9d9")
        self.Label15.configure(highlightcolor="black")
        self.Label15.configure(text='''Patient Information''')


        self.Button1 = tk.Button(self.Frame1)
        self.Button1.place(relx=0.776, rely=0.883, height=24, width=49)
        self.Button1.configure(activebackground="#ececec")
        self.Button1.configure(activeforeground="#000000")
        self.Button1.configure(background="#d9d9d9")
        self.Button1.configure(disabledforeground="#a3a3a3")
        self.Button1.configure(foreground="#000000")
        self.Button1.configure(highlightbackground="#d9d9d9")
        self.Button1.configure(highlightcolor="black")
        self.Button1.configure(pady="0")
        self.Button1.configure(text='''Submit''')
        self.Button1.configure(command = self.next_step)
        self.Button1.configure(command = self.submit)
        

        self.Button2 = tk.Button(self.Frame1)
        self.Button2.place(relx=0.09, rely=0.883, height=24, width=47)
        self.Button2.configure(activebackground="#ececec")
        self.Button2.configure(activeforeground="#000000")
        self.Button2.configure(background="#d9d9d9")
        self.Button2.configure(disabledforeground="#a3a3a3")
        self.Button2.configure(foreground="#000000")
        self.Button2.configure(highlightbackground="#d9d9d9")
        self.Button2.configure(highlightcolor="black")
        self.Button2.configure(pady="0")
        self.Button2.configure(text='''Cancel''')
        self.Button2.configure(command = cancel)


        self.Entry1_7 = tk.Entry(self.Frame1, validate="key" )
        self.Entry1_7.place(relx=0.418, rely=0.096, height=21, relwidth=0.504)
        self.Entry1_7.configure(background="white")
        self.Entry1_7.configure(disabledforeground="#a3a3a3")
        self.Entry1_7.configure(font="TkFixedFont")
        self.Entry1_7.configure(foreground="#000000")
        self.Entry1_7.configure(insertbackground="black")
        self.Entry1_7['validatecommand'] = (self.Entry1_7.register(testVal),'%P','%d')
        self.Entry1_7.configure(textvariable = entry_text5)


        self.Entry1 = tk.Entry(self.Frame1, validate="key" )
        self.Entry1.place(relx=0.418, rely=0.156, height=20, relwidth=0.504)
        self.Entry1.configure(background="white")
        self.Entry1.configure(disabledforeground="#a3a3a3")
        self.Entry1.configure(font="TkFixedFont")
        self.Entry1.configure(foreground="#000000")
        self.Entry1.configure(insertbackground="black")
        self.Entry1['validatecommand'] = (self.Entry1.register(testVal),'%P','%d')
        self.Entry1.configure(textvariable = entry_text3)
        

        self.Entry1_1 = tk.Entry(self.Frame1, validate="key" )
        self.Entry1_1.place(relx=0.418, rely=0.220, height=21, relwidth=0.504)
        self.Entry1_1.configure(background="white")
        self.Entry1_1.configure(disabledforeground="#a3a3a3")
        self.Entry1_1.configure(font="TkFixedFont")
        self.Entry1_1.configure(foreground="#000000")
        self.Entry1_1.configure(highlightbackground="#d9d9d9")
        self.Entry1_1.configure(highlightcolor="black")
        self.Entry1_1.configure(insertbackground="black")
        self.Entry1_1.configure(selectbackground="blue")
        self.Entry1_1.configure(selectforeground="white")
        #self.Entry1_1['validatecommand'] = (self.Entry1_1.register(testVal),'%P','%d')
        self.Entry1_1.insert(0, str(datetime.datetime.now()))
        self.Entry1_1.configure(state='disabled')
       
        

        self.Entry1_2 = tk.Entry(self.Frame1, validate="key")
        self.Entry1_2.place(relx=0.418, rely=0.290, height=21, relwidth=0.504)
        self.Entry1_2.configure(background="white")
        self.Entry1_2.configure(disabledforeground="#a3a3a3")
        self.Entry1_2.configure(font="TkFixedFont")
        self.Entry1_2.configure(foreground="#000000")
        self.Entry1_2.configure(highlightbackground="#d9d9d9")
        self.Entry1_2.configure(highlightcolor="black")
        self.Entry1_2.configure(insertbackground="black")
        self.Entry1_2.configure(selectbackground="blue")
        self.Entry1_2.configure(selectforeground="white")
        #self.Entry1_2['validatecommand'] = (self.Entry1_2.register(testVal1),'%P','%d')
        


        self.Entry1_3 = tk.Entry(self.Frame1, validate="key")
        self.Entry1_3.place(relx=0.418, rely=0.370, height=21, relwidth=0.504)
        self.Entry1_3.configure(background="white")
        self.Entry1_3.configure(disabledforeground="#a3a3a3")
        self.Entry1_3.configure(font="TkFixedFont")
        self.Entry1_3.configure(foreground="#000000")
        self.Entry1_3.configure(highlightbackground="#d9d9d9")
        self.Entry1_3.configure(highlightcolor="black")
        self.Entry1_3.configure(insertbackground="black")
        self.Entry1_3.configure(selectbackground="blue")
        self.Entry1_3.configure(selectforeground="white")
        self.Entry1_3['validatecommand'] = (self.Entry1_3.register(testVal),'%P','%d')
        self.Entry1_3.configure(textvariable = entry_text)
        
        self.Radiobutton1 = tk.Radiobutton(self.Frame1)
        self.Radiobutton1.place(relx=0.354, rely=0.442, relheight=0.065
                , relwidth=0.188)
        #self.Radiobutton1.configure(activebackground="#ececec")
        self.Radiobutton1.configure(activebackground="#d9d9d9")
        self.Radiobutton1.configure(activeforeground="#000000")
        self.Radiobutton1.configure(background="#d9d9d9")
        self.Radiobutton1.configure(disabledforeground="#a3a3a3")
        self.Radiobutton1.configure(foreground="#000000")
        self.Radiobutton1.configure(highlightbackground="#d9d9d9")
        self.Radiobutton1.configure(highlightcolor="black")
        self.Radiobutton1.configure(justify='left')
        self.Radiobutton1.configure(text='''Male''')
        self.Radiobutton1.configure(variable=self.var,value=1)
        #self.Radiobutton1.configure(value=1)

 
        self.Radiobutton2 = tk.Radiobutton(self.Frame1)
        self.Radiobutton2.place(relx=0.525, rely=0.442, relheight=0.065
                , relwidth=0.215)
        self.Radiobutton2.configure(activebackground="#ececec")
        self.Radiobutton2.configure(activeforeground="#000000")
        self.Radiobutton2.configure(background="#d9d9d9")
        self.Radiobutton2.configure(disabledforeground="#a3a3a3")
        self.Radiobutton2.configure(foreground="#000000")
        self.Radiobutton2.configure(highlightbackground="#d9d9d9")
        self.Radiobutton2.configure(highlightcolor="black")
        self.Radiobutton2.configure(justify='left')
        self.Radiobutton2.configure(text='''Female''')
        self.Radiobutton2.configure(variable=self.var,value=2)



        self.Radiobutton3 = tk.Radiobutton(self.Frame1)
        self.Radiobutton3.place(relx=0.756, rely=0.442, relheight=0.065
                , relwidth=0.173)
        self.Radiobutton3.configure(activebackground="#ececec")
        self.Radiobutton3.configure(activeforeground="#000000")
        self.Radiobutton3.configure(background="#d9d9d9")
        self.Radiobutton3.configure(disabledforeground="#a3a3a3")
        self.Radiobutton3.configure(foreground="#000000")
        self.Radiobutton3.configure(highlightbackground="#d9d9d9")
        self.Radiobutton3.configure(highlightcolor="black")
        self.Radiobutton3.configure(justify='left')
        self.Radiobutton3.configure(text='''Other''')
        self.Radiobutton3.configure(variable=self.var,value=3)

        self.Entry1_4 = tk.Entry(self.Frame1, validate="key")
        self.Entry1_4.place(relx=0.418, rely=0.525, height=21, relwidth=0.504)
        self.Entry1_4.configure(background="white")
        self.Entry1_4.configure(disabledforeground="#a3a3a3")
        self.Entry1_4.configure(font="TkFixedFont")
        self.Entry1_4.configure(foreground="#000000")
        self.Entry1_4.configure(highlightbackground="#d9d9d9")
        self.Entry1_4.configure(highlightcolor="black")
        self.Entry1_4.configure(insertbackground="black")
        self.Entry1_4.configure(selectbackground="blue")
        self.Entry1_4.configure(selectforeground="white")
        self.Entry1_4['validatecommand'] = (self.Entry1_4.register(testVal),'%P','%d')
        self.Entry1_4.configure(textvariable = entry_text1)

        
        self.Entry1_5 = tk.Entry(self.Frame1, validate="key")
        self.Entry1_5.place(relx=0.418, rely=0.604, height=21, relwidth=0.504)
        self.Entry1_5.configure(background="white")
        self.Entry1_5.configure(disabledforeground="#a3a3a3")
        self.Entry1_5.configure(font="TkFixedFont")
        self.Entry1_5.configure(foreground="#000000")
        self.Entry1_5.configure(highlightbackground="#d9d9d9")
        self.Entry1_5.configure(highlightcolor="black")
        self.Entry1_5.configure(insertbackground="black")
        self.Entry1_5.configure(selectbackground="blue")
        self.Entry1_5.configure(selectforeground="white")
        self.Entry1_5['validatecommand'] = (self.Entry1_5.register(testVal),'%P','%d')
        self.Entry1_5.configure(textvariable = entry_text2)

        self.Entry1_6 = tk.Entry(self.Frame1)
        self.Entry1_6.place(relx=0.418, rely=0.685, height=21, relwidth=0.504)
        self.Entry1_6.configure(background="white")
        self.Entry1_6.configure(disabledforeground="#a3a3a3")
        self.Entry1_6.configure(font="TkFixedFont")
        self.Entry1_6.configure(foreground="#000000")
        self.Entry1_6.configure(highlightbackground="#d9d9d9")
        self.Entry1_6.configure(highlightcolor="black")
        self.Entry1_6.configure(insertbackground="black")
        self.Entry1_6.configure(selectbackground="blue")
        self.Entry1_6.configure(selectforeground="white")
        #self.Entry1_6['validatecommand'] = (self.Entry1_6.register(testVal),'%P','%d')
        
        
        
        self.Entry1_8 = tk.Entry(self.Frame1)
        self.Entry1_8.place(relx=0.418, rely=0.760, height=21, relwidth=0.504)
        self.Entry1_8.configure(background="white")
        self.Entry1_8.configure(disabledforeground="#a3a3a3")
        self.Entry1_8.configure(font="TkFixedFont")
        self.Entry1_8.configure(foreground="#000000")
        self.Entry1_8.configure(highlightbackground="#d9d9d9")
        self.Entry1_8.configure(highlightcolor="black")
        self.Entry1_8.configure(insertbackground="black")
        self.Entry1_8.configure(selectbackground="blue")
        self.Entry1_8.configure(selectforeground="white")
        #self.Entry1_6['validatecommand'] = (self.Entry1_6.register(testVal),'%P','%d')
        
        
if __name__ == '__main__':
    vp_start_gui()






